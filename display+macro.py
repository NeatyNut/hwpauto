from pkgutil import extend_path
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
from openpyxl import load_workbook
import win32com.client as win32
import webbrowser
from tkinterdnd2 import *

up_collist = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y','Z']
# low_collist = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']
shtli = []

def click(event) :
    fd_path = fd_txt.get("1.0",'end-1c')
    hf_name = hp_txt.get("1.0",'end-1c')
    ex_name = ex_txt.get("1.0",'end-1c')
    ex_sh_name_index = ex_combo.current()
    ex_cls_col = ex_cls_txt.get("1.0",'end-1c')
    ex_uni_col = ex_uni_txt.get("1.0",'end-1c')
    ex_jeon_col = ex_jeon_txt.get("1.0",'end-1c')
    ex_cls_col_index = 0
    ex_uni_col_index = 0
    ex_jeon_col_index = 0

    ex_cls_col = ex_cls_col.replace("\t","")
    ex_cls_col = ex_cls_col.strip()

    ex_uni_col = ex_uni_col.replace("\t","")
    ex_uni_col = ex_uni_col.strip()
    
    ex_jeon_col = ex_jeon_col.replace("\t","")
    ex_jeon_col = ex_jeon_col.strip()

    #수록구분 열번호 얻기
    if ex_cls_col.isdigit() :
        ex_cls_col_index = ex_cls_col
    else :
        ex_cls_col = ex_cls_col.upper()

        for i in range(len(ex_cls_col)) :
            ex_cls_col_index = ex_cls_col_index + (26**(len(ex_cls_col)-i-1))*(up_collist.index(ex_cls_col[i])+1)

    #대학명 열번호 얻기
    if ex_uni_col.isdigit() :
        ex_uni_col_index = ex_uni_col
    else :
        ex_uni_col = ex_uni_col.upper()

        for i2 in range(len(ex_uni_col)) :
            ex_uni_col_index = ex_uni_col_index + (26**(len(ex_uni_col)-i2-1))*(up_collist.index(ex_uni_col[i2])+1)
        
    #전형명 열번호 얻기
    if ex_jeon_col.isdigit() :
        ex_jeon_col_index = ex_jeon_col
    else :
        ex_jeon_col = ex_jeon_col.upper()         

        for i3 in range(len(ex_jeon_col)) :
            ex_jeon_col_index = ex_jeon_col_index + (26**(len(ex_jeon_col)-i3-1))*(up_collist.index(ex_jeon_col[i3])+1)
        
    ex_cls_col_index = int(ex_cls_col_index)
    ex_uni_col_index = int(ex_uni_col_index)
    ex_jeon_col_index = int(ex_jeon_col_index)

    if fd_path.find("/") == -1 or hf_name.find(".hwp") == -1 or ex_name.find(".") == -1 or ex_sh_name_index == "" or ex_cls_col_index == 0 or ex_uni_col_index == 0 or ex_jeon_col_index == 0 :
        messagebox.showwarning("경고", "올바르게 입력해주세요")
        return

    def open(s) :
        # Automation 레지스트리 적용 필요
        # https://employeecoding.tistory.com/67
        hwp.RegisterModule("FilePathCheckDLL", "FilePathCheckerModule")
        hwp.Open(s)
        hwp.XHwpWindows.Item(0).Visible = True  # 실행된 한글 보이게 함.

    # 내용 입력
    def write(s) :
        hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
        hwp.HParameterSet.HInsertText.Text = s
        hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)

    # 폰트변경
    def font(s) :
        hwp.HAction.GetDefault("CharShape", hwp.HParameterSet.HCharShape.HSet)
        hwp.HParameterSet.HCharShape.FaceNameUser = s  # 글자모양 - 글꼴종류
        hwp.HParameterSet.HCharShape.FaceNameSymbol = s  # 글자모양 - 글꼴종류
        hwp.HParameterSet.HCharShape.FaceNameOther = s  # 글자모양 - 글꼴종류
        hwp.HParameterSet.HCharShape.FaceNameJapanese = s  # 글자모양 - 글꼴종류
        hwp.HParameterSet.HCharShape.FaceNameHanja = s  # 글자모양 - 글꼴종류
        hwp.HParameterSet.HCharShape.FaceNameLatin = s  # 글자모양 - 글꼴종류
        hwp.HParameterSet.HCharShape.FaceNameHangul = s  # 글자모양 - 글꼴종류

        hwp.HParameterSet.HCharShape.FontTypeUser = hwp.FontType("TTF")  # ?? 글자모양 - 폰트타입
        hwp.HParameterSet.HCharShape.FontTypeSymbol = hwp.FontType("TTF")  # 특수문자 글자모양 - 폰트타입
        hwp.HParameterSet.HCharShape.FontTypeOther = hwp.FontType("TTF")  # 일반 글자모양 - 폰트타입
        hwp.HParameterSet.HCharShape.FontTypeJapanese = hwp.FontType("TTF")  # 일본어 글자모양 - 폰트타입
        hwp.HParameterSet.HCharShape.FontTypeHanja = hwp.FontType("TTF")  # 한자 글자모양 - 폰트타입
        hwp.HParameterSet.HCharShape.FontTypeLatin = hwp.FontType("TTF")  # 라틴 글자모양 - 폰트타입
        hwp.HParameterSet.HCharShape.FontTypeHangul = hwp.FontType("TTF")  # 한글 글자모양 - 폰트타입
        hwp.HParameterSet.HCharShape.RatioSymbol = 100  #특수문자 장평
        hwp.HParameterSet.HCharShape.RatioHanja = 100  #한자 장평
        hwp.HParameterSet.HCharShape.SpacingLatin = 0 #라틴어 자간
        hwp.HParameterSet.HCharShape.RatioLatin = 100 #라틴어 장평
        hwp.HParameterSet.HCharShape.SpacingHangul = 0 #한글 자간
        hwp.HParameterSet.HCharShape.RatioHangul = 100 #한글 장평
        hwp.HAction.Execute("CharShape", hwp.HParameterSet.HCharShape.HSet)

    # 글자크기
    def fontsize(i) :
        hwp.HAction.GetDefault("CharShape", hwp.HParameterSet.HCharShape.HSet)
        hwp.HParameterSet.HCharShape.Height = hwp.PointToHwpUnit(i) # 글자모양 - 글자크기
        hwp.HAction.Execute("CharShape", hwp.HParameterSet.HCharShape.HSet)

    # 글자색깔
    def fontcolor(a, b, c) :
        hwp.HAction.GetDefault("CharShape", hwp.HParameterSet.HCharShape.HSet)
        hwp.HParameterSet.HCharShape.TextColor = hwp.RGBColor(a, b, c)
        hwp.HAction.Execute("CharShape", hwp.HParameterSet.HCharShape.HSet)

    # 표 만들기(줄, 열, 너비)
    def tlb_m(a, b, c) :
        hwp.HAction.GetDefault("TableCreate", hwp.HParameterSet.HTableCreation.HSet)
        hwp.HParameterSet.HTableCreation.Rows = a
        hwp.HParameterSet.HTableCreation.Cols = b
        hwp.HParameterSet.HTableCreation.WidthType = 2  # 너비 지정(0:단에맞춤, 1:문단에맞춤, 2:임의값)
        hwp.HParameterSet.HTableCreation.HeightType = 0  # 높이 지정(0:자동, 1:임의값)
        hwp.HParameterSet.HTableCreation.WidthValue = hwp.MiliToHwpUnit(178)  # 표 너비
        hwp.HParameterSet.HTableCreation.TableProperties.TreatAsChar = 1 # 표 글자 취급
        hwp.HAction.Execute("TableCreate", hwp.HParameterSet.HTableCreation.HSet)

        hwp.HAction.Run("TableCellBlock")
        hwp.HAction.Run("TableCellBlockExtend")
        hwp.HAction.Run("TableCellBlockExtend")

        hwp.HAction.GetDefault("TablePropertyDialog", hwp.HParameterSet.HShapeObject.HSet)
        hwp.HParameterSet.HShapeObject.HSet.SetItem("ShapeType", 3)
        hwp.HParameterSet.HShapeObject.HSet.SetItem("ShapeCellSize", 1)
        hwp.HParameterSet.HShapeObject.ShapeTableCell.Width = hwp.MiliToHwpUnit(c/b)
        hwp.HAction.Execute("TablePropertyDialog", hwp.HParameterSet.HShapeObject.HSet)

        hwp.HAction.Run("Cancel")

    # 셀 한줄 블록선택
    def tlb_sr(a) :
        hwp.HAction.Run("TableCellBlock")
        hwp.HAction.Run("TableCellBlockExtend")

        for i in range(a-1) :
            hwp.HAction.Run("TableRightCell")

    # 셀 한열 블록선택
    def tlb_sc(a) :
        hwp.HAction.Run("TableCellBlock")
        hwp.HAction.Run("TableCellBlockExtend")

        for i in range(a-1) :
            hwp.HAction.Run("TableLowerCell")

    # 테이블 전체 블럭선택
    def tlb_sa() :
        hwp.HAction.Run("TableCellBlock")
        hwp.HAction.Run("TableCellBlockExtend")
        hwp.HAction.Run("TableCellBlockExtend")

    # 한줄 선택
    def row_s() :
        hwp.HAction.Run("Select")
        hwp.HAction.Run("Select")
        hwp.HAction.Run("Select")
        hwp.HAction.Run("MoveLeft")

    def head_row() :
        hwp.HAction.GetDefault("TablePropertyDialog", hwp.HParameterSet.HShapeObject.HSet)
        hwp.HParameterSet.HShapeObject.HSet.SetItem("ShapeType", 3)
        hwp.HParameterSet.HShapeObject.HSet.SetItem("ShapeCellSize", 0)
        hwp.HParameterSet.HShapeObject.ShapeTableCell.Height = hwp.MiliToHwpUnit(4.5)
        hwp.HParameterSet.HShapeObject.ShapeTableCell.Header = 1
        hwp.HAction.Execute("TablePropertyDialog", hwp.HParameterSet.HShapeObject.HSet)

    # 테이블 배경색 지정 - RGB
    def tlb_c(a,b,c) :
        hwp.HAction.GetDefault("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)
        hwp.HParameterSet.HCellBorderFill.FillAttr.type = hwp.BrushType("NullBrush|WinBrush")
        hwp.HParameterSet.HCellBorderFill.FillAttr.WinBrushFaceColor = hwp.RGBColor(a, b, c)
        hwp.HParameterSet.HCellBorderFill.FillAttr.WinBrushHatchColor = hwp.RGBColor(a, b, c)
        hwp.HParameterSet.HCellBorderFill.FillAttr.WinBrushFaceStyle = hwp.HatchStyle("None")
        hwp.HParameterSet.HCellBorderFill.FillAttr.WindowsBrush = 1
        hwp.HAction.Execute("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)

    # 테이블 라인 정리
    def tlb_ln() :
        hwp.HAction.GetDefault("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)
        hwp.HParameterSet.HCellBorderFill.BorderTypeRight = hwp.HwpLineType("None")
        hwp.HParameterSet.HCellBorderFill.BorderTypeLeft = hwp.HwpLineType("None")
        hwp.HParameterSet.HCellBorderFill.BorderWidthBottom = hwp.HwpLineWidth("0.4mm")
        hwp.HParameterSet.HCellBorderFill.BorderWidthTop = hwp.HwpLineWidth("0.4mm")
        hwp.HAction.Execute("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)

    # 테이블 크기 정리
    def tlb_cols(a) :
        hwp.HAction.GetDefault("TablePropertyDialog", hwp.HParameterSet.HShapeObject.HSet)
        hwp.HParameterSet.HShapeObject.HSet.SetItem("ShapeType", 3)
        hwp.HParameterSet.HShapeObject.HSet.SetItem("ShapeCellSize", 1)
        hwp.HParameterSet.HShapeObject.ShapeTableCell.Width = hwp.MiliToHwpUnit(27.0)
        hwp.HAction.Execute("TablePropertyDialog", hwp.HParameterSet.HShapeObject.HSet)

    # 다른 이름으로 저장
    def fl_svas(name) :
        hwp.SaveAs(name,"HWP")
        hwp.XHwpWindows.Item(0).Visible = True  # 실행된 한글 보이게 함.

    # 테이블 각주 넣기
    def tlb_tem(tem) :
        hwp.HAction.Run("ShapeObjCaption")
        hwp.HAction.Run("SelectAll")
        hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
        hwp.HParameterSet.HInsertText.Text = tem
        hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)

    #1 엑셀내용 읽기
    wb = load_workbook(ex_name, data_only=True)   #원본파일
    shtli2 = wb.sheetnames
    ws = wb[shtli2[ex_sh_name_index]]
    if ex_cls_col_index > ws.max_column or ex_uni_col_index > ws.max_column or ex_jeon_col_index > ws.max_column :
        messagebox.showwarning("경고", "수록구분열/학교명열/표제목열 또는 엑셀파일을 확인해주세요")
        return

    if str(ws.cell(row=2, column = ex_cls_col_index).value).find("권") == -1 and str(ws.cell(row=2, column = ex_cls_col_index).value).find("PDF") == -1 : 
        answer0 = messagebox.askokcancel(
                title = "경고",
                message = "수록구분열인 " + str(ex_cls_col) + "열의 첫번째 내용 = " + str(ws.cell(row=2, column = ex_cls_col_index).value) + '\n' + "그대로 진행할까요?",
                icon = messagebox.WARNING)
        if answer0 != True :
            return

    if str(ws.cell(row=2, column = ex_uni_col_index).value).find("대학교") == -1 :
        answer1 = messagebox.askokcancel(
                title = "경고",
                message = "학교명열인 " + str(ex_uni_col) + "열의 첫번째 내용 = " + str(ws.cell(row=2, column = ex_uni_col_index).value) + '\n' + "그대로 진행할까요?",
                icon = messagebox.WARNING)
        if answer1 != True :
            return
    if str(ws.cell(row=2, column = ex_jeon_col_index).value).find("■") == -1 :
        answer2 = messagebox.askokcancel(
                title = "경고",
                message = "표제목열인 " + str(ex_jeon_col) + "열의 첫번째 내용파트 = " + str(ws.cell(row=2, column = ex_jeon_col_index).value) + '\n' + "그대로 진행할까요?",
                icon = messagebox.WARNING)
        if answer2 != True :
            return
    if ws.cell(row=1, column=ws.max_column).value != "자연" or ws.cell(row=1, column=ws.max_column-1).value != "인문" or str(ws.cell(row=2, column=ws.max_column-2).value).find("~") == -1 :
        answer3 = messagebox.askokcancel(
                title = "경고",
                message = "점수파트 헤드구성 = " + str(ws.cell(row=1, column=ws.max_column-2).value) + " " + str(ws.cell(row=1, column=ws.max_column-1).value) + " " + str(ws.cell(row=1, column=ws.max_column).value) + '\n' + "그대로 진행할까요?",
                icon = messagebox.WARNING)
        if answer3 != True :
            return
    
    rownum = 1
    rownum_max = ws.max_row

    # 해당 서식 파일 열기
    hwp = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")
    # hwp = win32.dynamic.Dispatch("HWPFrame.HwpObject")
    # hwp 열기 오류 시 https://martinii.fun/183 첫 번째 방법
    open(hf_name)
    #헤드 정리(첫 헤드)
    uni_name = ws.cell(row=rownum+1, column = ex_uni_col_index).value           ##학교명
    hwp.SetPos(9,0,0)  # 헤드 위치
    hwp.HAction.Run("SelectAll")
    write(uni_name) # 대학명 입력
    jeon_s = 2
    place = 4
    tb_place = 19
    tb_place_2 = 0

    while rownum -1 <= rownum_max :
        rownum += 1
        if ws.cell(row=rownum, column = ex_uni_col_index).value != ws.cell(row=rownum-1, column = ex_uni_col_index).value and rownum-1 != 1 : ##(좌표)대학명 비교
            fl_svas(fd_path + "/" + ws.cell(row=rownum-1, column = ex_cls_col_index).value + "_" 
                + ws.cell(row=rownum-1, column = ex_uni_col_index).value.replace("대학교", "대").replace("여자대","여대").replace("외국어대","외대").replace("교육대","교대").replace("과학기술원","과기원")
                + "_수시 수험자료집(입시결과).hwp")
            hwp.Quit()
            if rownum < rownum_max :
                hwp = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")
                # hwp = win32.dynamic.Dispatch("HWPFrame.HwpObject")
                open(hf_name)
                #헤드 정리(차 후 헤드)
                uni_name = ws.cell(row=rownum, column = ex_uni_col_index).value           ##학교명
                hwp.SetPos(9,0,0)  # 헤드 위치
                hwp.HAction.Run("SelectAll")
                write(uni_name) # 대학명 입력
                place = 4
                tb_place = 19
                # standard = "(대):대교협 “어디가” 발표 기준"           ##기준
                # hwp.SetPos(17,0,0)
                # hwp.HAction.Run("SelectAll")
                # write(standard) #(대):대교협 “어디가” 발표 기준, (학):대학별 발표 자료 (표):실제 합격자 표본조사 토대 성적
                # hwp.SetPos(18,0,24) #각주 변경 

            # 첫 전형이 한줄일 때
            if ws.cell(row=rownum, column=ex_jeon_col_index).value != ws.cell(row=rownum+1, column=ex_jeon_col_index).value : ##(좌표)위아래 전형 비교
                jeon_e = rownum  ######
                tb_row = jeon_e - jeon_s + 1
                information = ws.cell(row=jeon_s, column=ex_jeon_col_index).value      ##(좌표)전형명
                hwp.SetPos(0,place,0)
                write(information)
                hwp.SelectText(place, 2, place, len(information)+1)
                hwp.HAction.Run("CharShapeBold")

                # 표 헤드
                hwp.SetPos(0,place+1,0)

                write("[고교 내신 성적]")
                hwp.SelectText(place+1, 0, place+1, len("[고교 내신 성적]")+1)
                hwp.HAction.Run("CharShapeBold")

                hwp.SetPos(0,place+2,0)
                hwp.HAction.Run("BreakPara")
                hwp.SetPos(0,place+2,0)

                #표 만들기
                row = tb_row + 1
                col = 3
                tlb_place_plus = row*col
                tlb_m(row,col,190.30)   #표너비

                hwp.SetPos(tb_place,0,0)
                tlb_sr(col)
                tlb_c(102,102,102)
                fontcolor(255,255,255)
                head_row()
                hwp.HAction.Run("CharShapeBold")
                tlb_ln()

                hwp.SetPos(tb_place,0,0)
                hwp.HAction.Run("MoveDown")
                tlb_sc(row-1)
                for n1 in range(33) :
                    hwp.HAction.Run("TableResizeLineLeft")
                tlb_c(229,229,229)

                hwp.SetPos(tb_place,0,0)
                write("내신등급대")
                hwp.HAction.Run("TableRightCell")
                write("인문")
                hwp.HAction.Run("TableRightCell")
                write("자연")

                hwp.SetPos(tb_place,0,0)
                hwp.HAction.Run("TableRightCell")
                hwp.HAction.Run("TableCellBlock")
                hwp.HAction.Run("TableCellBlockExtend")
                hwp.HAction.Run("TableRightCell")
                for n2 in range(row-1) :
                    hwp.HAction.Run("TableLowerCell")

                hwp.HAction.Run("TableDistributeCellWidth")

                # 내용입력
                hwp.SetPos(tb_place,0,0)
                hwp.HAction.Run("TableLowerCell")
                n3 = 0
                for v in range(jeon_s, jeon_e+1) :
                    for h in range(3) :
                        hwp.SetPos(tb_place+3+n3,0,0)
                        write(ws.cell(row=v,column=ws.max_column-2+h).value)
                        n3 += 1
                #표 안 폰트 지정
                tlb_sa()
                tlb_ln()
                font("맑은 고딕")
                fontsize(7)
                hwp.HAction.Run("ParagraphShapeAlignCenter")

                place += 4
                jeon_s = jeon_e + 1
                tb_place += tlb_place_plus

                hwp.SetPos(0,place,0)
                hwp.HAction.Run("BreakPara")
                hwp.HAction.Run("BreakPara")
                hwp.HAction.Run("BreakPara")
                hwp.SetPos(0,place,0)

        else :
            # 이 외 표만들기
            if ws.cell(row=rownum, column=ex_jeon_col_index).value != ws.cell(row=rownum+1, column=ex_jeon_col_index).value or ws.cell(row=rownum, column=ex_uni_col_index).value != ws.cell(row=rownum+1, column=ex_uni_col_index).value : #위아래 전형 비교
                jeon_e = rownum  ######
                tb_row = jeon_e - jeon_s + 1
                information = ws.cell(row=jeon_s, column=ex_jeon_col_index).value      ##(좌표)전형명
                hwp.SetPos(0,place,0)
                write(information)
                hwp.SelectText(place, 2, place, len(information)+1)
                hwp.HAction.Run("CharShapeBold")

                # 표 헤드
                hwp.SetPos(0,place+1,0)

                write("[고교 내신 성적]")
                hwp.SelectText(place+1, 0, place+1, len("[고교 내신 성적]")+1)
                hwp.HAction.Run("CharShapeBold")

                hwp.SetPos(0,place+2,0)
                hwp.HAction.Run("BreakPara")
                hwp.SetPos(0,place+2,0)

                #표 만들기
                row = tb_row + 1
                col = 3
                tlb_place_plus = row*col
                tlb_m(row,col,190.30)   #표너비

                hwp.SetPos(tb_place,0,0)
                tlb_sr(col)
                tlb_c(102,102,102)
                fontcolor(255,255,255)
                head_row()
                hwp.HAction.Run("CharShapeBold")
                tlb_ln()

                hwp.SetPos(tb_place,0,0)
                hwp.HAction.Run("MoveDown")
                tlb_sc(row-1)
                for n1 in range(33) :
                    hwp.HAction.Run("TableResizeLineLeft")
                tlb_c(229,229,229)

                hwp.SetPos(tb_place,0,0)
                write("내신등급대")
                hwp.HAction.Run("TableRightCell")
                write("인문")
                hwp.HAction.Run("TableRightCell")
                write("자연")

                hwp.SetPos(tb_place,0,0)
                hwp.HAction.Run("TableRightCell")
                hwp.HAction.Run("TableCellBlock")
                hwp.HAction.Run("TableCellBlockExtend")
                hwp.HAction.Run("TableRightCell")
                for n2 in range(row-1) :
                    hwp.HAction.Run("TableLowerCell")

                hwp.HAction.Run("TableDistributeCellWidth")

                # 내용입력
                hwp.SetPos(tb_place,0,0)
                hwp.HAction.Run("TableLowerCell")
                n3 = 0
                for v in range(jeon_s, jeon_e+1) :
                    for h in range(3) :
                        hwp.SetPos(tb_place+3+n3,0,0)
                        write(ws.cell(row=v,column=ws.max_column-2+h).value)
                        n3 += 1

                #표 안 폰트지정
                tlb_sa()
                tlb_ln()
                font("맑은 고딕")
                fontsize(7)
                hwp.HAction.Run("ParagraphShapeAlignCenter")

                place += 4
                jeon_s = jeon_e + 1
                tb_place += tlb_place_plus

                hwp.SetPos(0,place,0)
                hwp.HAction.Run("BreakPara")
                hwp.HAction.Run("BreakPara")
                hwp.HAction.Run("BreakPara")
                hwp.SetPos(0,place,0)

    messagebox.showinfo("완료","완료하였습니다")
    gui.destroy()
    return("break")

#경로 선택 시(저장하는 곳)
def fd(event) :
   fd_txt.configure(state=NORMAL)
   fd_txt.delete('1.0', END)       
   file_folder = filedialog.askdirectory(initialdir="/",\
                  title = "저장할 경로를 지정해주세요")
   if file_folder == '':
      messagebox.showwarning("경고", "경로를 지정해주세요")    #파일 선택 안했을 때 메세지 출력
   else :
      fd_txt.insert(1.0, file_folder)
      fd_txt.configure(state=DISABLED)
   return("break")

# 파일 드래그앤 드랍
def fddrag(event) :
    if event.data.find(".") == -1 :  
        file_folder = event.data.replace("{","").replace("}","")  # 양괄호 삭제
        fd_txt.configure(state=NORMAL)
        fd_txt.delete('1.0', END)
        fd_txt.insert(1.0, file_folder)
        fd_txt.configure(state=DISABLED)
        return("break")
    else :
        messagebox.showwarning("경고", "폴더를 추가 하세요")
        return("break")

#한글 선택 시(양식)
def hp(event) :
   hp_txt.configure(state=NORMAL)
   hp_txt.delete('1.0', END)
   file_hwp = filedialog.askopenfilename(initialdir="/",\
                  title = "양식파일(한글)을 선택 해 주세요",\
                     filetype = [("한/글파일","*.hwp *.hwpx")])
   if file_hwp == '':
      messagebox.showwarning("경고", "파일을 추가 하세요")    #파일 선택 안했을 때 메세지 출력
   else :
      hp_txt.insert(1.0, file_hwp)
      hp_txt.configure(state=DISABLED)
   return("break")

# 한글 파일 드래그앤 드랍
def hpdrag(event) :
    file_hwp = event.data
    file_hwp = file_hwp.replace("{","").replace("}","")  # 양괄호 삭제
    if file_hwp.endswith(".hwp") or file_hwp.endswith(".hwpx") :  # 한글은 괄호가 들어옴
        hp_txt.configure(state=NORMAL)
        hp_txt.delete('1.0', END)
        hp_txt.insert(1.0, file_hwp)
        hp_txt.configure(state=DISABLED)
        return("break")
    else :
        messagebox.showwarning("경고", "한글파일을 추가 하세요")
        return("break")

#엑셀 선택 시
def ex(event) :
   ex_txt.configure(state=NORMAL)
   ex_txt.delete('1.0', END)
   file_ex = filedialog.askopenfilename(initialdir="/",\
                  title = "원본DB파일(엑셀)을 선택 해 주세요",\
                     filetypes = (("*.xlsx","*xlsx"),("*.xls","*xls")))
   if file_ex == '':
      messagebox.showwarning("경고", "파일을 추가 하세요")    #파일 선택 안했을 때 메세지 출력
   else :
      wb = load_workbook(file_ex, data_only=True)
      shtli = wb.sheetnames
      ex_txt.insert(1.0, file_ex)
      ex_txt.configure(state=DISABLED)
      ex_combo['values'] = shtli

   return("break")

#엑셀 드래그앤 드랍
def exdrag(event) :
    file_ex = event.data
    file_ex = file_ex.replace("{","").replace("}","")  # 양괄호 삭제
    if file_ex.endswith(".xlsx") or file_ex.endswith(".xls") :
        wb = load_workbook(file_ex, data_only=True)
        shtli = wb.sheetnames
        ex_txt.configure(state=NORMAL)
        ex_txt.delete('1.0', END)
        ex_txt.insert(1.0, file_ex)
        ex_txt.configure(state=DISABLED)
        ex_combo['values'] = shtli
        return("break")
    else :
        messagebox.showwarning("경고", "엑셀파일을 추가 하세요")
        return("break")

#종료 원할때
def stop(event) :
    if messagebox.askokcancel("종료", "종료하고 싶습니까?"):
        gui.destroy()
    return("break")

#url클릭
def callback(url):
    webbrowser.open_new(url)
#tab버튼
def focus_next_widget(event):
    event.widget.tk_focusNext().focus()
    return("break")
#shift+tab버튼
def focus_prev_widget(event):
    event.widget.tk_focusPrev().focus()
    return("break")

gui = Tk()
gui.title("수시 수험자료집 입시결과 매크로")
gui.geometry("430x450")

Label(gui, text="DB선택(엑셀)", font=('돋움', 10, 'bold')).place(x=1,y=5)
ex_txt = Text(gui, height=2, width=50)
ex_txt.place(x=1,y=30)
ex_txt.bind("<Tab>", focus_next_widget)
ex_txt.bind("<Shift-Tab>", focus_prev_widget)
ex_txt.drop_target_register(DND_FILES)
ex_txt.dnd_bind("<<Drop>>",exdrag)
ex_txt.configure(state=DISABLED, bg='light yellow2')
ex_but = Button(gui, text="찾아보기", font=('돋움', 10), cursor="hand2")
ex_but.bind('<Button-1>',ex)
ex_but.bind('<Return>',ex)
ex_but.place(x=360,y=32)


Label(gui, text="시트선택", font=('돋움', 10, 'bold')).place(x=1,y=80)
ex_combo = ttk.Combobox(gui, values = shtli, cursor="hand2")
ex_combo.place(x=1, y=100)
Label(gui, text="※시트 미선택 시 첫번째 시트 자동선택됨",font=('돋움', 8),fg='blue').place(x=1, y=120)
Label(gui, text="열번호(1/A/a) 입력", font=('돋움', 10, 'bold')).place(x=220, y=67)
Label(gui, text="1.수록구분    2. 학교명    3. 표제목", font=('돋움', 8)).place(x=220, y=82)

ex_cls_txt = Text(gui, height=1, width=7)
ex_cls_txt.place(x=220, y=100)
ex_cls_txt.bind("<Tab>", focus_next_widget)
ex_cls_txt.bind("<Shift-Tab>", focus_prev_widget)

ex_uni_txt = Text(gui, height=1, width=7)
ex_uni_txt.place(x=285, y=100)
ex_uni_txt.bind("<Tab>", focus_next_widget)
ex_uni_txt.bind("<Shift-Tab>", focus_prev_widget)

ex_jeon_txt = Text(gui, height=1, width=7)
ex_jeon_txt.place(x=350, y=100)
ex_jeon_txt.bind("<Tab>", focus_next_widget)
ex_jeon_txt.bind("<Shift-Tab>", focus_prev_widget)

Label(gui, text="양식선택(hwp)", font=('돋움', 10, 'bold')).place(x=1, y=145)
hp_txt = Text(gui, height=2, width=50)
hp_txt.place(x=1,y=165)
hp_txt.bind("<Tab>", focus_next_widget)
hp_txt.bind("<Shift-Tab>", focus_prev_widget)
hp_txt.configure(state=DISABLED, bg='light yellow2')
hp_txt.drop_target_register(DND_FILES)
hp_txt.dnd_bind("<<Drop>>",hpdrag)

hp_but = Button(gui, text="찾아보기", font=('돋움', 10), cursor="hand2")
hp_but.bind('<Button-1>',hp)
hp_but.bind('<Return>',hp)
hp_but.place(x=360,y=167)
Label(gui, text="※한글양식은 새로 만들지 말고, 기존 것 연도 수정 후 사용할 것",font=('돋움', 8),fg='red').place(x=1, y=195)

Label(gui, text="저장할 경로 지정", font=('돋움', 10, 'bold')).place(x=1, y=220)
fd_txt = Text(gui, height=2, width=50)
fd_txt.place(x=1, y=240)
fd_txt.bind("<Tab>", focus_next_widget)
fd_txt.bind("<Shift-Tab>", focus_prev_widget)
fd_txt.configure(state=DISABLED, bg='light yellow2')
fd_txt.drop_target_register(DND_FILES)
fd_txt.dnd_bind("<<Drop>>",fddrag)


fd_but = Button(gui, text="찾아보기", font=('돋움', 10), cursor="hand2")
fd_but.bind('<Button-1>',fd)
fd_but.bind('<Return>',fd)
fd_but.place(x=360,y=242)


action_but = Button(gui, text="실행하기", font=('돋움', 10, 'bold'), fg='white', cursor="hand2")
action_but.bind('<Button-1>',click)
action_but.bind('<Return>',click)

action_but.place(x=360,y=300)
action_but.config(bg="blue")

Label(gui, text="주의사항",font=('돋움', 10, 'bold'),fg='red').place(x=1, y=330)
Label(gui, text="1. 엑셀 정렬(학교명, 전형유형, 전형명, 점수구간 순)",font=('돋움', 10),fg='blue').place(x=1, y=350)
Label(gui, text="2. 엑셀파일 [첫번째줄=헤드, 두번째줄 이후=내용]으로 구성할 것",font=('돋움', 10),fg='blue').place(x=1, y=370)
Label(gui, text="3. 엑셀파일의 작업내용 맨 끝열은 점수파트",font=('돋움', 10)).place(x=1, y=390)
Label(gui, text="4. 한개 이상의 한글파일 실행 중일 시, 오류발생함",font=('돋움', 10,'bold'),fg='red').place(x=1, y=410)
Label(gui, text="#. 강제종료를 원할 땐 작업관리자를 이용할 것",font=('돋움', 10,'bold')).place(x=1, y=430)
Label(gui, text="made by kim", font=('맑은 고딕', 7),fg='red').place(x=360, y=400)

web = Label(gui, text="[필수]한글자동화를 위한 레지스트리 적용 시(클릭)",font=('돋움', 10, 'bold'),fg='white', cursor="hand2")
web.config(bg="black")
web.place(x=1, y=300)
web.bind("<Button-1>", lambda e: callback("https://employeecoding.tistory.com/67"))

exit_but = Button(gui, text="종료하기", font=('돋움', 10), cursor="hand2")
exit_but.place(x=360,y=420)
exit_but.bind('<Button-1>',stop)
exit_but.bind('<Return>',stop)

#hp_txt는 한글파일 경로, ex_txt는 엑셀파일 경로, ex_uni_txt는 대학명 열번호, ex_jeon_txt 전형명 열번호, 구간은 끝열 -2, 인문은 끝열 -1, 자연은 끝열
gui.mainloop()
