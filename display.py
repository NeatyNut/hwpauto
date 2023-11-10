from pkgutil import extend_path
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
from openpyxl import load_workbook
import macro

shtli = []
def click() :
   try :
      macro
   except :
      messagebox.showwarning("경고", "올바르게 입력해주세요")

#경로 선택 시(저장하는 곳)
def fd() :
   file_folder = filedialog.askdirectory(initialdir="/",\
                  title = "저장할 경로를 지정해주세요")
   if file_folder == '':
      messagebox.showwarning("경고", "경로를 지정해주세요")    #파일 선택 안했을 때 메세지 출력
   else :
      fd_txt.insert(1.0, file_folder)

#한글 선택 시(양식)
def hp() :
   file_hwp = filedialog.askopenfilename(initialdir="/",\
                  title = "양식파일(한글)을 선택 해 주세요",\
                     filetype = [("한/글파일","*.hwp *.hwpx")])
   if file_hwp == '':
      messagebox.showwarning("경고", "파일을 추가 하세요")    #파일 선택 안했을 때 메세지 출력
   else :
      hp_txt.insert(1.0, file_hwp)

#엑셀 선택 시
def ex() :
   file_ex = filedialog.askopenfilename(initialdir="/",\
                  title = "원본DB파일(엑셀)을 선택 해 주세요",\
                     filetypes = (("*.xlsx","*xlsx"),("*.xls","*xls")))
   if file_ex == '':
      messagebox.showwarning("경고", "파일을 추가 하세요")    #파일 선택 안했을 때 메세지 출력
   else :
      wb = load_workbook(file_ex, data_only=True)
      shtli = wb.sheetnames
      ex_txt.insert(1.0, file_ex)
      ex_combo['values'] = shtli

gui = Tk()
gui.title("수시 수험자료집 입시결과 매크로")
gui.geometry("430x360")

Label(gui, text="DB선택(엑셀)", font=('돋움', 10, 'bold')).place(x=1,y=5)
ex_txt = Text(gui, height=2, width=50)
ex_txt.place(x=1,y=30)
ex_but = Button(gui, text="찾아보기", font=('돋움', 10), command=lambda:ex())
ex_but.place(x=360,y=32)

Label(gui, text="시트선택", font=('돋움', 10, 'bold')).place(x=1,y=80)
ex_combo = ttk.Combobox(gui, values = shtli)
ex_combo.place(x=1, y=100)

Label(gui, text="대학명(열번호)", font=('돋움', 10, 'bold')).place(x=200, y=80)
Label(gui, text="전형명(열번호)", font=('돋움', 10, 'bold')).place(x=320, y=80)
ex_uni_txt = Text(gui, height=1, width=12)
ex_uni_txt.place(x=200, y=100)
ex_jeon_txt = Text(gui, height=1, width=12)
ex_jeon_txt.place(x=320, y=100)

Label(gui, text="양식선택(한글)", font=('돋움', 10, 'bold')).place(x=1, y=140)
hp_txt = Text(gui, height=2, width=50)
hp_txt.place(x=1,y=160)
hp_but = Button(gui, text="찾아보기", font=('돋움', 10), command=lambda:hp())
hp_but.place(x=360,y=162)

Label(gui, text="저장할 경로 지정", font=('돋움', 10, 'bold')).place(x=1, y=220)
fd_txt = Text(gui, height=2, width=50)
fd_txt.place(x=1, y=240)
fd_but = Button(gui, text="찾아보기", font=('돋움', 10), command=lambda:fd())
fd_but.place(x=360,y=242)

action_but = Button(gui, text="실행하기", font=('돋움', 10), command=lambda:click())
action_but.place(x=360,y=300)

Label(gui, text="한군데라도 미선택 시 실행되지 않습니다  =>",font=('돋움', 11, 'bold'),fg='red').place(x=1, y=300)

#hp_txt는 한글파일 경로, ex_txt는 엑셀파일 경로, ex_uni_txt는 대학명 열번호, ex_jeon_txt 전형명 열번호, 구간은 끝열 -2, 인문은 끝열 -1, 자연은 끝열
gui.mainloop()

