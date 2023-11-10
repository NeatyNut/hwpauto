import win32com.client as win32
import datetime
from display import *

up_collist = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y','Z']
# low_collist = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']

fd_path = fd_txt.get("1.0",'end-1c')
hf_name = hp_txt.get("1.0",'end-1c')
ex_name = ex_txt.get("1.0",'end-1c')
ex_sh_name_index = ex_combo.current()
ex_uni_col = ex_uni_txt.get("1.0",'end-1c')
ex_jeon_col = ex_jeon_txt.get("1.0",'end-1c')
ex_uni_col_index = 0
ex_jeon_col_index = 0

#대학명 열번호 얻기
if ex_uni_col.isdigit() :
    ex_uni_col_index = ex_uni_col
else :
    ex_uni_col = ex_uni_col.upper()

    for i in range(len(ex_uni_col)) :
        ex_uni_col_index = ex_uni_col_index + (26**(len(ex_uni_col)-i-1))*(up_collist.index(ex_uni_col[i])+1)

#전형명 열번호 얻기
if ex_jeon_col.isdigit() :
    ex_jeon_col_index = ex_jeon_col
else :
    ex_jeon_col = ex_jeon_col.upper()         

    for i2 in range(len(ex_jeon_col)) :
        ex_jeon_col_index = ex_jeon_col_index + (26**(len(ex_jeon_col)-i2-1))*(up_collist.index(ex_jeon_col[i2])+1)

currentDateTime = datetime.datetime.now()
date = currentDateTime.date()
year = str(int(date.strftime("%Y")) + 1)

def open(s) :
    # Automation 레지스트리 적용 필요
    # https://employeecoding.tistory.com/67
    hwp.RegisterModule("FilePathCheckDLL", "FilePathCheckerModule")
    hwp.Open(s)
    hwp.XHwpWindows.Item(0).Visible = True  # 실행된 한글 보이게 함.

# 내용 입력
def write(s) :
    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
    hwp.HParameterSet.HInsertText.Text = s
    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)

# 폰트변경
def font(s) :
    hwp.HAction.GetDefault("CharShape", hwp.HParameterSet.HCharShape.HSet)
    hwp.HParameterSet.HCharShape.FaceNameUser = s  # 글자모양 - 글꼴종류
    hwp.HParameterSet.HCharShape.FaceNameSymbol = s  # 글자모양 - 글꼴종류
    hwp.HParameterSet.HCharShape.FaceNameOther = s  # 글자모양 - 글꼴종류
    hwp.HParameterSet.HCharShape.FaceNameJapanese = s  # 글자모양 - 글꼴종류
    hwp.HParameterSet.HCharShape.FaceNameHanja = s  # 글자모양 - 글꼴종류
    hwp.HParameterSet.HCharShape.FaceNameLatin = s  # 글자모양 - 글꼴종류
    hwp.HParameterSet.HCharShape.FaceNameHangul = s  # 글자모양 - 글꼴종류

    hwp.HParameterSet.HCharShape.FontTypeUser = hwp.FontType("TTF")  # 글자모양 - 폰트타입
    hwp.HParameterSet.HCharShape.FontTypeSymbol = hwp.FontType("TTF")  # 글자모양 - 폰트타입
    hwp.HParameterSet.HCharShape.FontTypeOther = hwp.FontType("TTF")  # 글자모양 - 폰트타입
    hwp.HParameterSet.HCharShape.FontTypeJapanese = hwp.FontType("TTF")  # 글자모양 - 폰트타입
    hwp.HParameterSet.HCharShape.FontTypeHanja = hwp.FontType("TTF")  # 글자모양 - 폰트타입
    hwp.HParameterSet.HCharShape.FontTypeLatin = hwp.FontType("TTF")  # 글자모양 - 폰트타입
    hwp.HParameterSet.HCharShape.FontTypeHangul = hwp.FontType("TTF")  # 글자모양 - 폰트타입
    hwp.HAction.Execute("CharShape", hwp.HParameterSet.HCharShape.HSet)

# 글자크기
def fontsize(i) :
    hwp.HAction.GetDefault("CharShape", hwp.HParameterSet.HCharShape.HSet)
    hwp.HParameterSet.HCharShape.Height = hwp.PointToHwpUnit(i) # 글자모양 - 글자크기
    hwp.HAction.Execute("CharShape", hwp.HParameterSet.HCharShape.HSet)

# 글자색깔
def fontcolor(a, b, c) :
    hwp.HAction.GetDefault("CharShape", hwp.HParameterSet.HCharShape.HSet)
    hwp.HParameterSet.HCharShape.TextColor = hwp.RGBColor(a, b, c)
    hwp.HAction.Execute("CharShape", hwp.HParameterSet.HCharShape.HSet)

# 표 만들기(줄, 열, 너비)
def tlb_m(a, b, c) :
    hwp.HAction.GetDefault("TableCreate", hwp.HParameterSet.HTableCreation.HSet)
    hwp.HParameterSet.HTableCreation.Rows = a
    hwp.HParameterSet.HTableCreation.Cols = b
    hwp.HParameterSet.HTableCreation.WidthType = 2  # 너비 지정(0:단에맞춤, 1:문단에맞춤, 2:임의값)
    hwp.HParameterSet.HTableCreation.HeightType = 0  # 높이 지정(0:자동, 1:임의값)
    hwp.HParameterSet.HTableCreation.WidthValue = hwp.MiliToHwpUnit(178)  # 표 너비
    hwp.HParameterSet.HTableCreation.TableProperties.TreatAsChar = 1 # 표 글자 취급
    hwp.HAction.Execute("TableCreate", hwp.HParameterSet.HTableCreation.HSet)

    hwp.HAction.Run("TableCellBlock")
    hwp.HAction.Run("TableCellBlockExtend")
    hwp.HAction.Run("TableCellBlockExtend")

    hwp.HAction.GetDefault("TablePropertyDialog", hwp.HParameterSet.HShapeObject.HSet)
    hwp.HParameterSet.HShapeObject.HSet.SetItem("ShapeType", 3)
    hwp.HParameterSet.HShapeObject.HSet.SetItem("ShapeCellSize", 1)
    hwp.HParameterSet.HShapeObject.ShapeTableCell.Width = hwp.MiliToHwpUnit(c/b)
    hwp.HAction.Execute("TablePropertyDialog", hwp.HParameterSet.HShapeObject.HSet)

    hwp.HAction.Run("Cancel")

# 셀 한줄 블록선택
def tlb_sr(a) :
    hwp.HAction.Run("TableCellBlock")
    hwp.HAction.Run("TableCellBlockExtend")

    for i in range(a-1) :
        hwp.HAction.Run("TableRightCell")

# 셀 한열 블록선택
def tlb_sc(a) :
    hwp.HAction.Run("TableCellBlock")
    hwp.HAction.Run("TableCellBlockExtend")

    for i in range(a-1) :
        hwp.HAction.Run("TableLowerCell")

# 테이블 전체 블럭선택
def tlb_sa() :
    hwp.HAction.Run("TableCellBlock")
    hwp.HAction.Run("TableCellBlockExtend")
    hwp.HAction.Run("TableCellBlockExtend")

# 한줄 선택
def row_s() :
    hwp.HAction.Run("Select")
    hwp.HAction.Run("Select")
    hwp.HAction.Run("Select")
    hwp.HAction.Run("MoveLeft")

# 테이블 배경색 지정 - RGB
def tlb_c(a,b,c) :
    hwp.HAction.GetDefault("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)
    hwp.HParameterSet.HCellBorderFill.FillAttr.type = hwp.BrushType("NullBrush|WinBrush")
    hwp.HParameterSet.HCellBorderFill.FillAttr.WinBrushFaceColor = hwp.RGBColor(a, b, c)
    hwp.HParameterSet.HCellBorderFill.FillAttr.WinBrushHatchColor = hwp.RGBColor(a, b, c)
    hwp.HParameterSet.HCellBorderFill.FillAttr.WinBrushFaceStyle = hwp.HatchStyle("None")
    hwp.HParameterSet.HCellBorderFill.FillAttr.WindowsBrush = 1
    hwp.HAction.Execute("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)

# 테이블 라인 정리
def tlb_ln() :
    hwp.HAction.GetDefault("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)
    hwp.HParameterSet.HCellBorderFill.BorderTypeRight = hwp.HwpLineType("None")
    hwp.HParameterSet.HCellBorderFill.BorderTypeLeft = hwp.HwpLineType("None")
    hwp.HParameterSet.HCellBorderFill.BorderWidthBottom = hwp.HwpLineWidth("0.4mm")
    hwp.HParameterSet.HCellBorderFill.BorderWidthTop = hwp.HwpLineWidth("0.4mm")
    hwp.HAction.Execute("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)

# 테이블 크기 정리
def tlb_cols(a) :
    hwp.HAction.GetDefault("TablePropertyDialog", hwp.HParameterSet.HShapeObject.HSet)
    hwp.HParameterSet.HShapeObject.HSet.SetItem("ShapeType", 3)
    hwp.HParameterSet.HShapeObject.HSet.SetItem("ShapeCellSize", 1)
    hwp.HParameterSet.HShapeObject.ShapeTableCell.Width = hwp.MiliToHwpUnit(27.0)
    hwp.HAction.Execute("TablePropertyDialog", hwp.HParameterSet.HShapeObject.HSet)

# 다른 이름으로 저장
def fl_svas(name) :
    hwp.SaveAs(name,"HWP")
    hwp.XHwpWindows.Item(0).Visible = True  # 실행된 한글 보이게 함.

# 테이블 각주 넣기
def tlb_tem(tem) :
    hwp.HAction.Run("ShapeObjCaption")
    hwp.HAction.Run("SelectAll")
    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
    hwp.HParameterSet.HInsertText.Text = tem
    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)

#1 엑셀내용 읽기

from openpyxl import load_workbook
try :
    wb = load_workbook(ex_name, data_only=True)   #원본파일
except :
    print(ex_name)
    exit()
shtli2 = wb.sheetnames
ws = wb[shtli2[ex_sh_name_index]]
rownum = 1
rownum_max = ws.max_row

# 해당 서식 파일 열기
hwp = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")
# hwp = win32.dynamic.Dispatch("HWPFrame.HwpObject")
# hwp 열기 오류 시 https://martinii.fun/183 첫 번째 방법
open(hf_name)
#헤드 정리(첫 헤드)
uni_name = ws.cell(row=rownum+1, column = ex_uni_col_index).value           ##학교명
hwp.SetPos(9,0,0)  # 헤드 위치
hwp.HAction.Run("SelectAll")
write(uni_name) # 대학명 입력
jeon_s = 2
place = 4
tb_place = 19
tb_place_2 = 0

while rownum -1 != rownum_max :
    rownum += 1
    if ws.cell(row=rownum, column = ex_uni_col_index).value != ws.cell(row=rownum-1, column = ex_uni_col_index).value and rownum-1 != 1 : ##(좌표)대학명 비교
        fl_svas(fd_path + "/" + year + ws.cell(row=rownum-1, column = ex_uni_col_index).value + "_수시 수험자료집(입시결과).hwp")
        hwp.Quit()
        if rownum != rownum_max :
            hwp = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")
            # hwp = win32.dynamic.Dispatch("HWPFrame.HwpObject")
            open(hf_name)
            #헤드 정리(차 후 헤드)
            uni_name = ws.cell(row=rownum, column = ex_uni_col_index).value           ##학교명
            hwp.SetPos(9,0,0)  # 헤드 위치
            hwp.HAction.Run("SelectAll")
            write(uni_name) # 대학명 입력
            place = 4
            tb_place = 19
        else :
            fl_svas(fd_path + "/" + year + ws.cell(row=rownum-1, column = ex_uni_col_index).value + "_수시 수험자료집(입시결과).hwp")
            hwp.Quit()
            # standard = "(대):대교협 “어디가” 발표 기준"           ##기준
            # hwp.SetPos(17,0,0)
            # hwp.HAction.Run("SelectAll")
            # write(standard) #(대):대교협 “어디가” 발표 기준, (학):대학별 발표 자료 (표):실제 합격자 표본조사 토대 성적
            # hwp.SetPos(18,0,24) #각주 변경 

        # 첫 전형이 한줄일 때
        if ws.cell(row=rownum, column=ex_jeon_col_index).value != ws.cell(row=rownum+1, column=ex_jeon_col_index).value : ##(좌표)위아래 전형 비교
            jeon_e = rownum  ######
            tb_row = jeon_e - jeon_s + 1
            information = ws.cell(row=jeon_s, column=11).value      ##(좌표)전형명
            hwp.SetPos(0,place,0)
            write(information)
            hwp.SelectText(place, 2, place, len(information)+1)
            hwp.HAction.Run("CharShapeBold")

            # 표 헤드
            hwp.SetPos(0,place+1,0)

            write("[고교 내신 성적]")
            hwp.SelectText(place+1, 0, place+1, len("[고교 내신 성적]")+1)
            hwp.HAction.Run("CharShapeBold")

            hwp.SetPos(0,place+2,0)
            hwp.HAction.Run("BreakPara")
            hwp.SetPos(0,place+2,0)

            #표 만들기
            row = tb_row + 1
            col = 3
            tlb_place_plus = row*col
            tlb_m(row,col,190.30)   #표너비
            tlb_sa()
            tlb_ln()
            font("맑은 고딕")
            fontsize(7)
            hwp.HAction.Run("ParagraphShapeAlignCenter")

            hwp.SetPos(tb_place,0,0)
            tlb_sr(col)
            tlb_c(102,102,102)
            fontcolor(255,255,255)
            hwp.HAction.Run("CharShapeBold")
            tlb_ln()

            hwp.SetPos(tb_place,0,0)
            hwp.HAction.Run("MoveDown")
            tlb_sc(row-1)
            for n1 in range(33) :
                hwp.HAction.Run("TableResizeLineLeft")
            tlb_c(229,229,229)

            hwp.SetPos(tb_place,0,0)
            write("내신등급대")
            hwp.HAction.Run("TableRightCell")
            write("인문")
            hwp.HAction.Run("TableRightCell")
            write("자연")

            hwp.SetPos(tb_place,0,0)
            hwp.HAction.Run("TableRightCell")
            hwp.HAction.Run("TableCellBlock")
            hwp.HAction.Run("TableCellBlockExtend")
            hwp.HAction.Run("TableRightCell")
            for n2 in range(row-1) :
                hwp.HAction.Run("TableLowerCell")

            hwp.HAction.Run("TableDistributeCellWidth")

            # 내용입력
            hwp.SetPos(tb_place,0,0)
            hwp.HAction.Run("TableLowerCell")
            n3 = 0
            for v in range(jeon_s, jeon_e+1) :
                for h in range(3) :
                    hwp.SetPos(tb_place+3+n3,0,0)
                    write(ws.cell(row=v,column=12+h).value)
                    n3 += 1

            place += 4
            jeon_s = jeon_e + 1
            tb_place += tlb_place_plus

            hwp.SetPos(0,place,0)
            hwp.HAction.Run("BreakPara")
            hwp.HAction.Run("BreakPara")
            hwp.HAction.Run("BreakPara")
            hwp.SetPos(0,place,0)

    else :
        # 이 외 표만들기
        if ws.cell(row=rownum, column=ex_jeon_col_index).value != ws.cell(row=rownum+1, column=ex_jeon_col_index).value : #위아래 전형 비교
            jeon_e = rownum  ######
            tb_row = jeon_e - jeon_s + 1
            information = ws.cell(row=jeon_s, column=11).value      ##(좌표)전형명
            hwp.SetPos(0,place,0)
            write(information)
            hwp.SelectText(place, 2, place, len(information)+1)
            hwp.HAction.Run("CharShapeBold")

            # 표 헤드
            hwp.SetPos(0,place+1,0)

            write("[고교 내신 성적]")
            hwp.SelectText(place+1, 0, place+1, len("[고교 내신 성적]")+1)
            hwp.HAction.Run("CharShapeBold")

            hwp.SetPos(0,place+2,0)
            hwp.HAction.Run("BreakPara")
            hwp.SetPos(0,place+2,0)

            #표 만들기
            row = tb_row + 1
            col = 3
            tlb_place_plus = row*col
            tlb_m(row,col,190.30)   #표너비
            tlb_sa()
            tlb_ln()
            font("맑은 고딕")
            fontsize(7)
            hwp.HAction.Run("ParagraphShapeAlignCenter")

            hwp.SetPos(tb_place,0,0)
            tlb_sr(col)
            tlb_c(102,102,102)
            fontcolor(255,255,255)
            hwp.HAction.Run("CharShapeBold")
            tlb_ln()

            hwp.SetPos(tb_place,0,0)
            hwp.HAction.Run("MoveDown")
            tlb_sc(row-1)
            for n1 in range(33) :
                hwp.HAction.Run("TableResizeLineLeft")
            tlb_c(229,229,229)

            hwp.SetPos(tb_place,0,0)
            write("내신등급대")
            hwp.HAction.Run("TableRightCell")
            write("인문")
            hwp.HAction.Run("TableRightCell")
            write("자연")

            hwp.SetPos(tb_place,0,0)
            hwp.HAction.Run("TableRightCell")
            hwp.HAction.Run("TableCellBlock")
            hwp.HAction.Run("TableCellBlockExtend")
            hwp.HAction.Run("TableRightCell")
            for n2 in range(row-1) :
                hwp.HAction.Run("TableLowerCell")

            hwp.HAction.Run("TableDistributeCellWidth")

            # 내용입력
            hwp.SetPos(tb_place,0,0)
            hwp.HAction.Run("TableLowerCell")
            n3 = 0
            for v in range(jeon_s, jeon_e+1) :
                for h in range(3) :
                    hwp.SetPos(tb_place+3+n3,0,0)
                    write(ws.cell(row=v,column=12+h).value)
                    n3 += 1

            place += 4
            jeon_s = jeon_e + 1
            tb_place += tlb_place_plus

            hwp.SetPos(0,place,0)
            hwp.HAction.Run("BreakPara")
            hwp.HAction.Run("BreakPara")
            hwp.HAction.Run("BreakPara")
            hwp.SetPos(0,place,0)
